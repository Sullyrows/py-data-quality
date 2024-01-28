from openpyxl import Workbook
from openpyxl.styles import Alignment, Side, Border
from openpyxl.worksheet.table import Table
from importlib import resources 
from itertools import chain 
import json

def main_gen(): 
    """generate main dataset"""
    with resources.path("py_data_quality","example") as f: 
        test_data = f.parents[1] / "tests/data"
        my_path = test_data / "excel_tests.json"

        json_values = json.loads(my_path.read_text())
    
    # build workbook
    workbook = Workbook()

    # deterministically build worksheets 
    for _value_dict in json_values: 
        # get _values from value_dict 
        sheet_values = _value_dict.get("values")

        # create sheet and add values 
        new_sheet = workbook.create_sheet(_value_dict.get("sheet_name"))

        # use json to determine values 
        cell_array = [[(row+1, col+1, v) for col, v in enumerate(values)] for row, values in enumerate(sheet_values)]
        cell_array = list(chain.from_iterable(cell_array))

        [new_sheet.cell(row, col, value) for row,col,value in cell_array]

        # print(True)

        # check for modifications 
        modifications = _value_dict.get("modifications")
        if modifications is not None: 
            for mod in modifications: 
                
                cell = new_sheet.cell(mod.get("row"), mod.get("col"))

                # check for modification type 
                match mod.keys():
                    case _ if "indent" in mod.keys(): 
                        cell.alignment = Alignment(indent=mod.get("indent"))
                    case _ if "alignment" in mod.keys(): 
                        # alignment value
                        _align = mod.get("alignment")

                        # format values
                        kwarg = {_align : "center"}

                        # now determine the 
                        cell.alignnment = Alignment(**kwarg)
                    case _ if "border" in mod.keys(): 
                        # get nested values from border
                        _border_dict = mod.get("border")

                        # format style
                        _side = Side(
                            color=_border_dict.get("color"),
                            border_style = _border_dict.get("style")
                        )
                        # build border
                        _b_type = _border_dict.get("type")
                        match _b_type:
                            case "left": 
                                _border = Border(left = _side)
                            case "right": 
                                _border = Border(right = _side)
                            case "top": 
                                _border = Border(top = _side)
                            case "bottom": 
                                _border = Border(bottom = _side)
                            case _: 
                                raise NotImplementedError(_border)

                        # add cell border 
                        cell.border = _border

        table_dict = _value_dict.get("table")
        if table_dict is not None: 
            # define table structure
            for tbl in table_dict: 
                # define table 
                my_table = Table(**tbl)

                # add table to worksheet 
                new_sheet.add_table(my_table)


    # cleanup "Sheet" sheet if it exists 
    sheet_present = "Sheet" in workbook.sheetnames
    if sheet_present:
        _sheet = workbook["Sheet"]
        workbook.remove(_sheet)

    workbook.save(filename=test_data / "excel_cases.xlsx")


if __name__ == "__main__": 
    main_gen()